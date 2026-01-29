import streamlit as st
import pandas as pd
import openpyxl
import os
import base64
import re

# Page config
st.set_page_config(
    page_title="Good Sports Research Statistics",
    page_icon="📊",
    layout="wide"
)

# Function to convert image to base64
def get_base64_image(image_path):
    """Convert image to base64 for embedding in HTML"""
    try:
        with open(image_path, "rb") as img_file:
            return base64.b64encode(img_file.read()).decode()
    except:
        return None

# Try to load logo
logo_base64 = get_base64_image("good_sports_logo.png")

# Custom CSS with Good Sports actual colors
st.markdown("""
<style>
    /* Import Google Fonts */
    @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600;700&display=swap');
    
    /* Main app styling */
    .stApp {
        background: linear-gradient(135deg, #f5f7fa 0%, #ffffff 100%);
    }
    
    /* Custom header with logo */
    .custom-header {
        background: linear-gradient(135deg, #8B9F3E 0%, #A8B968 100%);
        padding: 2rem;
        border-radius: 15px;
        margin-bottom: 2rem;
        box-shadow: 0 4px 15px rgba(139, 159, 62, 0.3);
        display: flex;
        align-items: center;
        gap: 2rem;
    }
    
    .logo-container {
        flex-shrink: 0;
    }
    
    .logo-container img {
        height: 80px;
        background: white;
        padding: 10px 20px;
        border-radius: 10px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    
    .header-text {
        flex-grow: 1;
    }
    
    .custom-header h1 {
        color: white;
        font-family: 'Montserrat', sans-serif;
        font-weight: 700;
        font-size: 2.5rem;
        margin: 0;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.1);
    }
    
    .custom-header p {
        color: white;
        font-family: 'Montserrat', sans-serif;
        font-size: 1.1rem;
        margin-top: 0.5rem;
        opacity: 0.95;
    }
    
    /* Filter section */
    .filter-section {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        margin-bottom: 2rem;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        border-top: 4px solid #8B9F3E;
    }
    
    /* Dropdown styling with Good Sports green */
    .stSelectbox > div > div {
        background-color: white;
        border: 2px solid #8B9F3E;
        border-radius: 8px;
        font-family: 'Montserrat', sans-serif;
        font-weight: 600;
    }
    
    /* Stat cards */
    .stat-card {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        margin-bottom: 1.5rem;
        border-left: 5px solid #8B9F3E;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        transition: transform 0.2s, box-shadow 0.2s;
    }
    
    .stat-card:hover {
        transform: translateY(-3px);
        box-shadow: 0 4px 15px rgba(139, 159, 62, 0.2);
    }
    
    /* New data card - special styling */
    .stat-card.new-data {
        border-left: 5px solid #FF6B35;
        background: linear-gradient(to right, #FFF5F0 0%, white 10%);
    }
    
    /* Updated data card - special styling */
    .stat-card.updated-data {
        border-left: 5px solid #0066A1;
        background: linear-gradient(to right, #E8F4F8 0%, white 10%);
    }
    
    /* Replaced/old data card */
    .stat-card.replaced-data {
        border-left: 5px solid #95A5A6;
        background: linear-gradient(to right, #F5F5F5 0%, white 10%);
        opacity: 0.85;
    }
    
    .year-badge {
        background: linear-gradient(135deg, #0066A1 0%, #0082CC 100%);
        color: white;
        padding: 0.5rem 1rem;
        border-radius: 20px;
        font-weight: 700;
        font-size: 1.1rem;
        display: inline-block;
        margin-bottom: 1rem;
        font-family: 'Montserrat', sans-serif;
        box-shadow: 0 2px 5px rgba(0, 102, 161, 0.3);
    }
    
    /* Badge for new/updated/replaced */
    .new-badge {
        background: linear-gradient(135deg, #FF6B35 0%, #FF8C42 100%);
        color: white;
        padding: 0.4rem 0.9rem;
        border-radius: 15px;
        font-weight: 600;
        font-size: 0.85rem;
        display: inline-block;
        margin-left: 0.5rem;
        font-family: 'Montserrat', sans-serif;
    }
    
    .updated-badge {
        background: linear-gradient(135deg, #0066A1 0%, #0082CC 100%);
        color: white;
        padding: 0.4rem 0.9rem;
        border-radius: 15px;
        font-weight: 600;
        font-size: 0.85rem;
        display: inline-block;
        margin-left: 0.5rem;
        font-family: 'Montserrat', sans-serif;
    }
    
    .replaced-badge {
        background: linear-gradient(135deg, #95A5A6 0%, #7F8C8D 100%);
        color: white;
        padding: 0.4rem 0.9rem;
        border-radius: 15px;
        font-weight: 600;
        font-size: 0.85rem;
        display: inline-block;
        margin-left: 0.5rem;
        font-family: 'Montserrat', sans-serif;
    }
    
    .stat-text {
        color: #2C3E50;
        font-size: 1.05rem;
        line-height: 1.6;
        font-family: 'Montserrat', sans-serif;
        margin: 1rem 0;
    }
    
    .link-info {
        background: #E8F4F8;
        padding: 0.75rem 1rem;
        border-radius: 8px;
        margin: 1rem 0;
        font-size: 0.95rem;
        font-family: 'Montserrat', sans-serif;
        border-left: 3px solid #0066A1;
    }
    
    .link-info.replaced-link {
        background: #F5F5F5;
        border-left: 3px solid #95A5A6;
    }
    
    .internal-link {
        color: #0066A1;
        font-weight: 600;
        text-decoration: none;
        cursor: pointer;
    }
    
    .internal-link:hover {
        text-decoration: underline;
    }
    
    .read-more-btn {
        background: linear-gradient(135deg, #8B9F3E 0%, #A8B968 100%);
        color: white;
        padding: 0.6rem 1.5rem;
        border-radius: 25px;
        text-decoration: none;
        font-weight: 600;
        display: inline-block;
        font-family: 'Montserrat', sans-serif;
        transition: transform 0.2s, box-shadow 0.2s;
        box-shadow: 0 2px 8px rgba(139, 159, 62, 0.3);
    }
    
    .read-more-btn:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(139, 159, 62, 0.4);
        text-decoration: none;
        color: white;
        background: linear-gradient(135deg, #9BB04A 0%, #B8C777 100%);
    }
    
    .source-text {
        color: #7F8C8D;
        font-size: 0.9rem;
        font-family: 'Montserrat', sans-serif;
        margin-top: 0.5rem;
    }
    
    /* Category selector label */
    label {
        font-family: 'Montserrat', sans-serif;
        font-weight: 600;
        color: #2C3E50;
        font-size: 1.1rem;
    }
    
    /* Info boxes */
    .stInfo {
        background-color: #E8F4F8;
        border-left: 5px solid #0066A1;
        border-radius: 8px;
        font-family: 'Montserrat', sans-serif;
    }
    
    /* Footer */
    .footer {
        text-align: center;
        padding: 2rem;
        color: #7F8C8D;
        font-family: 'Montserrat', sans-serif;
        margin-top: 3rem;
    }
    
    /* Stats counter */
    .stats-counter {
        background: linear-gradient(135deg, #0066A1 0%, #0082CC 100%);
        color: white;
        padding: 0.75rem 1.5rem;
        border-radius: 25px;
        font-weight: 600;
        display: inline-block;
        margin-bottom: 1.5rem;
        font-family: 'Montserrat', sans-serif;
        box-shadow: 0 3px 10px rgba(0, 102, 161, 0.3);
    }
    
    /* Category title with accent */
    .category-title {
        color: #8B9F3E;
        font-family: 'Montserrat', sans-serif;
        font-weight: 700;
        border-bottom: 3px solid #8B9F3E;
        padding-bottom: 0.5rem;
        margin-bottom: 1.5rem;
    }
</style>
""", unsafe_allow_html=True)

# File path
EXCEL_FILE = '2025 Good Sports Research Project_Aidan Conte.xlsx'

# Load data and hyperlinks
@st.cache_data
def load_data():
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name='Research', engine='openpyxl')
        return df, None
    except Exception as e:
        return None, f"❌ Error loading file: {str(e)}"

@st.cache_resource
def load_workbook():
    """Load workbook to extract hyperlinks"""
    try:
        return openpyxl.load_workbook(EXCEL_FILE)
    except Exception as e:
        st.warning(f"Could not load hyperlinks: {e}")
        return None

def get_hyperlink(workbook, row_index):
    """Extract hyperlink from Excel cell in the Source column (column B)"""
    if workbook is None:
        return None
    try:
        sheet = workbook['Research']
        excel_row = row_index + 2
        cell = sheet.cell(row=excel_row, column=2)
        if cell.hyperlink:
            return cell.hyperlink.target
    except Exception as e:
        pass
    return None

def extract_row_numbers(text):
    """Extract row numbers from text like 'Yes (ROW 120, 121)' or 'Updated (ROW 4)'"""
    if pd.isna(text):
        return []
    text = str(text)
    # Find all numbers after "ROW"
    matches = re.findall(r'ROW\s*(\d+)', text, re.IGNORECASE)
    return [int(m) for m in matches]

def excel_row_to_pandas_index(excel_row):
    """Convert Excel row number (1-indexed with header) to pandas index (0-indexed)"""
    # Excel row 2 = pandas index 0 (row 1 is header)
    return excel_row - 2

def is_new_data(row):
    """Check if this is new data based on Priority column"""
    priority = row.get('Priority for Updated Stat', '')
    return str(priority).strip() == 'New Data'

def is_updated_data(row):
    """Check if this is updated data based on Stat updated column"""
    updated = row.get('Stat updated? (see comment)', '')
    priority = row.get('Priority for Updated Stat', '')
    # Check if priority says "Updated" or stat updated says "Yes"
    return (pd.notna(priority) and 'Updated' in str(priority)) or \
           (pd.notna(updated) and str(updated).strip().startswith('Yes'))

def is_replaced_data(row):
    """Check if this stat has been replaced by a newer one"""
    updated = row.get('Stat updated? (see comment)', '')
    return pd.notna(updated) and (str(updated).strip().startswith('Yes') or 'ROW' in str(updated))

def get_replacement_rows(row, df):
    """Get the rows that replaced this stat"""
    updated = row.get('Stat updated? (see comment)', '')
    excel_rows = extract_row_numbers(updated)
    pandas_indices = [excel_row_to_pandas_index(r) for r in excel_rows]
    # Filter to valid indices
    return [idx for idx in pandas_indices if 0 <= idx < len(df)]

def get_replaced_row(row):
    """Get the row that this stat replaces"""
    priority = row.get('Priority for Updated Stat', '')
    excel_rows = extract_row_numbers(priority)
    if excel_rows:
        pandas_idx = excel_row_to_pandas_index(excel_rows[0])
        return pandas_idx if pandas_idx >= 0 else None
    return None

# Custom header with logo
if logo_base64:
    header_html = f"""
    <div class="custom-header">
        <div class="logo-container">
            <img src="data:image/png;base64,{logo_base64}" alt="Good Sports Logo">
        </div>
        <div class="header-text">
            <h1>Research Statistics Database</h1>
            <p>Comprehensive data supporting youth sports and physical activity initiatives</p>
        </div>
    </div>
    """
else:
    header_html = """
    <div class="custom-header">
        <div class="header-text">
            <h1>📊 Good Sports Research Statistics</h1>
            <p>Comprehensive data supporting youth sports and physical activity initiatives</p>
        </div>
    </div>
    """

st.markdown(header_html, unsafe_allow_html=True)

# Load data
df, error = load_data()

if error:
    st.error(error)
    st.info(f"**Current working directory:** `{os.getcwd()}`")
    st.stop()

# Load workbook for hyperlinks
workbook = load_workbook()

# Get unique categories
categories = sorted([cat for cat in df['Category'].unique() if pd.notna(cat)])

# Filter section
st.markdown('<div class="filter-section">', unsafe_allow_html=True)
st.markdown("### 🔍 Filter Research Statistics")

col1, col2 = st.columns([2, 1])

with col1:
    selected_category = st.selectbox(
        "Select Category:",
        ["-- All Categories --"] + categories
    )

with col2:
    data_filter = st.selectbox(
        "Data Status:",
        ["All Data", "New Data Only", "Updated Data Only", "Show Replaced Data"]
    )

st.markdown('</div>', unsafe_allow_html=True)

# Apply filters
if selected_category == "-- All Categories --":
    filtered_df = df.copy()
    display_title = "All Statistics"
else:
    filtered_df = df[df['Category'] == selected_category].copy()
    display_title = selected_category

# Apply data status filter
if data_filter == "New Data Only":
    filtered_df = filtered_df[filtered_df.apply(is_new_data, axis=1)]
    display_title += " - New Data"
elif data_filter == "Updated Data Only":
    filtered_df = filtered_df[filtered_df.apply(is_updated_data, axis=1)]
    display_title += " - Updated Data"
elif data_filter == "Show Replaced Data":
    # Show only old stats that have been replaced
    filtered_df = filtered_df[filtered_df.apply(is_replaced_data, axis=1)]
    display_title += " - Replaced Data"
else:
    # For "All Data", hide replaced stats by default
    filtered_df = filtered_df[~filtered_df.apply(is_replaced_data, axis=1)]

# Display results
if len(filtered_df) == 0:
    st.info("No statistics found matching your filters.")
else:
    st.markdown(f'<h2 class="category-title">{display_title}</h2>', unsafe_allow_html=True)
    
    # Count stats with breakdown
    total = len(filtered_df)
    new_count = sum(filtered_df.apply(is_new_data, axis=1))
    updated_count = sum(filtered_df.apply(is_updated_data, axis=1))
    
    stats_info = f'<div class="stats-counter">📈 {total} statistic(s) found'
    if data_filter == "All Data":
        if new_count > 0:
            stats_info += f' • {new_count} new'
        if updated_count > 0:
            stats_info += f' • {updated_count} updated'
    stats_info += '</div>'
    
    st.markdown(stats_info, unsafe_allow_html=True)
    
    # Display each stat
    for idx, (orig_idx, row) in enumerate(filtered_df.iterrows(), 1):
        # Determine card styling and status
        is_new = is_new_data(row)
        is_updated = is_updated_data(row)
        is_replaced = is_replaced_data(row)
        
        card_class = "stat-card"
        if is_replaced:
            card_class += " replaced-data"
        elif is_new:
            card_class += " new-data"
        elif is_updated:
            card_class += " updated-data"
        
        # Create stat card
        year = row['Year'] if pd.notna(row['Year']) else "N/A"
        stat_text = row['Stat'] if pd.notna(row['Stat']) else "No description available"
        source = row['Source'] if pd.notna(row['Source']) else None
        
        # Get hyperlink from Excel
        hyperlink = get_hyperlink(workbook, orig_idx)
        
        # Build HTML for stat card
        card_html = f'<div class="{card_class}">'
        card_html += f'<div class="year-badge">📅 {year}</div>'
        
        # Add status badges
        if is_replaced:
            card_html += '<span class="replaced-badge">📦 REPLACED</span>'
        elif is_new:
            card_html += '<span class="new-badge">✨ NEW</span>'
        elif is_updated:
            card_html += '<span class="updated-badge">🔄 UPDATED</span>'
        
        card_html += f'<div class="stat-text">{stat_text}</div>'
        
        # Show links to replacement stats (for old/replaced stats)
        if is_replaced:
            replacement_indices = get_replacement_rows(row, df)
            if replacement_indices:
                card_html += '<div class="link-info">'
                card_html += '🔄 <strong>This stat has been replaced by:</strong><br>'
                for repl_idx in replacement_indices:
                    repl_year = df.iloc[repl_idx]['Year'] if pd.notna(df.iloc[repl_idx]['Year']) else "N/A"
                    card_html += f'• <span class="internal-link">Updated stat from {repl_year} (Row {repl_idx + 2})</span><br>'
                card_html += '</div>'
        
        # Show link to original stat (for updated stats)
        if is_updated and not is_replaced:
            orig_idx_link = get_replaced_row(row)
            if orig_idx_link is not None and orig_idx_link < len(df):
                orig_year = df.iloc[orig_idx_link]['Year'] if pd.notna(df.iloc[orig_idx_link]['Year']) else "N/A"
                card_html += '<div class="link-info replaced-link">'
                card_html += f'📌 <strong>This replaces:</strong> Original stat from {orig_year} (Row {orig_idx_link + 2})'
                card_html += '</div>'
        
        if source:
            card_html += f'<div class="source-text">📄 <strong>Source:</strong> {source}</div>'
        
        if hyperlink:
            card_html += f'<br><a href="{hyperlink}" target="_blank" class="read-more-btn">🔗 Read Full Article</a>'
        
        card_html += '</div>'
        
        st.markdown(card_html, unsafe_allow_html=True)

# Footer with Good Sports branding
st.markdown("""
<div class="footer">
    <strong style="color: #8B9F3E;">Good Sports Research Project</strong> | 2025<br>
    <em>Empowering youth through sports and physical activity</em>
</div>
""", unsafe_allow_html=True)
